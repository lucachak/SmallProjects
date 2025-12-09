
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
from tkinter import filedialog

class ExcelHandler:
    def __init__(self, email_handler, status_callback=None,
                 DATE_HEADER="VENCIMENTO DA LICENÇA",
                 EMAIL_HEADER="EMAIL"):

        self.email_handler = email_handler
        self.status_callback = status_callback  # function to update GUI status
        self.DATE_HEADER = DATE_HEADER
        self.EMAIL_HEADER = EMAIL_HEADER

    def update_status(self, text):
        if self.status_callback:
            self.status_callback(text)
        else:
            print(text)

    def find_column(self, ws, target_header):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val and str(val).strip().lower() == target_header.lower():
                return col
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val and target_header.lower() in str(val).strip().lower():
                return col
        return None

    def process_spreadsheet(self, file_path):
        if not os.path.exists(file_path):
            self.update_status("❌ File not found.")
            return

        wb = load_workbook(file_path)
        ws = wb.active

        date_col = self.find_column(ws, self.DATE_HEADER)
        email_col = self.find_column(ws, self.EMAIL_HEADER)

        if date_col is None:
            self.update_status(f"❌ '{self.DATE_HEADER}' column not found.")
            return

        fill_expired = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_warning = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        fill_safe    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_header  = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        today = datetime.today().date()
        one_month = today + timedelta(days=30)

        # style header
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill_header
            cell.font = Font(bold=True)
            cell.alignment = center_align

        # process rows
        for row in range(2, ws.max_row + 1):
            raw_date = ws.cell(row=row, column=date_col).value
            exp_date = None
            try:
                if isinstance(raw_date, (datetime, date)):
                    exp_date = raw_date.date() if isinstance(raw_date, datetime) else raw_date
                else:
                    exp_date = pd.to_datetime(str(raw_date), dayfirst=True, errors="coerce")
                    if pd.notna(exp_date):
                        exp_date = exp_date.date()
            except Exception:
                exp_date = None

            if not exp_date:
                continue

            if exp_date < today:
                fill = fill_expired
            elif today <= exp_date <= one_month:
                fill = fill_warning
                if email_col:
                    email_value = ws.cell(row=row, column=email_col).value
                    name_value = ws.cell(row=row, column=1).value
                    if email_value:
                        self.email_handler.send_email(email_value, name_value, exp_date)
            else:
                fill = fill_safe

            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=row, column=col)
                c.fill = fill
                c.alignment = center_align

        # auto column width
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                value = ws.cell(row=row, column=col).value
                if value is None:
                    continue
                text = value.strftime("%d/%m/%Y") if isinstance(value, (datetime, date)) else str(value)
                max_len = max(max_len, len(text))
            ws.column_dimensions[get_column_letter(col)].width = min((max_len + 2) * 1.2, 80)

        folder = os.path.dirname(file_path)
        filename = os.path.basename(file_path)
        new_file_path = os.path.join(folder, f"processed_{filename}")
        wb.save(new_file_path)
        self.update_status(f"✅ Saved: {new_file_path}")

    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx *.xlsm")]
        )
        if file_path:
            self.process_spreadsheet(file_path)

